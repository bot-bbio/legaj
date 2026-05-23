import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_tracker(file_path):
    if os.path.exists(file_path):
        print(f"File already exists: {file_path}")
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Applications"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = [
        "Company", "Job Title", "Location", "Date Applied", 
        "Link/URL", "Status", "Resume Used", "Cover Letter Used", "Notes"
    ]
    
    ws.append(headers)
    
    # Styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Steel Blue
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Adjust column widths
    column_widths = {
        "A": 20, # Company
        "B": 25, # Job Title
        "C": 15, # Location
        "D": 15, # Date Applied
        "E": 30, # Link
        "F": 15, # Status
        "G": 25, # Resume
        "H": 25, # Cover Letter
        "I": 40  # Notes
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    # Add a sample row with placeholders or instructions
    sample_row = [
        "Example Corp", 
        "Product Manager", 
        "New York, NY", 
        "2026-05-22", 
        "https://example.com/job", 
        "Applied", 
        "Roberto_Montero_Resume_Tailored.pdf", 
        "Example_Corp_Cover_Letter.pdf", 
        "Initial application sent. Follow up in 2 weeks."
    ]
    ws.append(sample_row)
    
    # Style sample row with thin border
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.border = thin_border
        if col_idx in [4, 6]:  # Date, Status
            cell.alignment = Alignment(horizontal="center")
            
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)
    print(f"Successfully created tracker at: {file_path}")

if __name__ == "__main__":
    import sys
    target = sys.argv[1] if len(sys.argv) > 1 else "C:\\Users\\molus\\projects\\legaj\\references\\job-tracker.xlsx"
    create_tracker(target)
