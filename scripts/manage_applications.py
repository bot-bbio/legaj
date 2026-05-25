import sys
import os
import datetime
import json

def migrate_if_needed(json_path):
    # Determine xlsx path from json path
    base, ext = os.path.splitext(json_path)
    xlsx_path = base + ".xlsx"
    
    if os.path.exists(json_path):
        return
        
    if os.path.exists(xlsx_path):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            ws = wb.active
            apps = []
            for row in range(2, ws.max_row + 1):
                comp = ws.cell(row=row, column=1).value
                role = ws.cell(row=row, column=2).value
                loc = ws.cell(row=row, column=3).value
                date = ws.cell(row=row, column=4).value
                link = ws.cell(row=row, column=5).value
                status = ws.cell(row=row, column=6).value
                res = ws.cell(row=row, column=7).value
                cl = ws.cell(row=row, column=8).value
                notes = ws.cell(row=row, column=9).value
                if comp or role:
                    apps.append({
                        "company": comp or "",
                        "role": role or "",
                        "location": loc or "",
                        "date": str(date) if date else "",
                        "link": link or "",
                        "status": status or "",
                        "resume": res or "",
                        "cover_letter": cl or "",
                        "notes": notes or ""
                    })
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(apps, f, indent=2)
            print(f"Migrated Excel tracker data from {xlsx_path} to {json_path}")
        except Exception as e:
            print(f"Migration warning: Could not read {xlsx_path}: {e}")
            # Write empty list
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump([], f, indent=2)
    else:
        # Create empty json tracker file
        os.makedirs(os.path.dirname(json_path), exist_ok=True)
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump([], f, indent=2)

def load_apps(file_path):
    migrate_if_needed(file_path)
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading applications from JSON: {e}")
        return []

def save_apps(file_path, apps):
    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(apps, f, indent=2)
    except Exception as e:
        print(f"Error saving applications to JSON: {e}")

def add_application(file_path, company, role, location, link, status="Applied", resume="", cover_letter="", notes=""):
    apps = load_apps(file_path)
    date_str = datetime.date.today().strftime("%Y-%m-%d")
    
    new_app = {
        "company": company,
        "role": role,
        "location": location,
        "date": date_str,
        "link": link,
        "status": status,
        "resume": resume,
        "cover_letter": cover_letter,
        "notes": notes
    }
    apps.append(new_app)
    save_apps(file_path, apps)
    print(f"Successfully added application for {role} at {company} to tracker.")

def update_application_status(file_path, company, role, new_status, notes=None):
    apps = load_apps(file_path)
    found = False
    for app in apps:
        if app.get("company", "").lower().strip() == company.lower().strip() and app.get("role", "").lower().strip() == role.lower().strip():
            app["status"] = new_status
            if notes:
                current_notes = app.get("notes", "")
                app["notes"] = f"{current_notes} | {notes}" if current_notes else notes
            found = True
            
    if found:
        save_apps(file_path, apps)
        print(f"Updated status for {role} at {company} to '{new_status}'.")
    else:
        print(f"No application found matching {role} at {company}.")

def delete_application(file_path, company, role):
    apps = load_apps(file_path)
    new_apps = []
    found = False
    for app in apps:
        if app.get("company", "").lower().strip() == company.lower().strip() and app.get("role", "").lower().strip() == role.lower().strip():
            found = True
        else:
            new_apps.append(app)
    if found:
        save_apps(file_path, new_apps)
        print(f"Successfully deleted application for {role} at {company}.")
    else:
        print(f"No application found matching {role} at {company}.")

def scan_emails_and_sync(file_path, email_addr, password, imap_server):
    import imaplib
    import email
    from email.header import decode_header
    
    print(f"Connecting to IMAP server {imap_server}...")
    try:
        mail = imaplib.IMAP4_SSL(imap_server)
        mail.login(email_addr, password)
        mail.select("inbox")
        
        status, messages = mail.search(None, '(OR OR BODY "application" BODY "interview" SUBJECT "application" SUBJECT "interview")')
        if status != "OK":
            print("No matching emails found.")
            return
            
        apps = load_apps(file_path)
        email_ids = messages[0].split()
        print(f"Scanning {len(email_ids)} emails for application status updates...")
        
        updates_made = False
        for e_id in email_ids[-20:]:
            res, msg_data = mail.fetch(e_id, "(RFC822)")
            if res != "OK":
                continue
                
            for response_part in msg_data:
                if isinstance(response_part, tuple):
                    msg = email.message_from_bytes(response_part[1])
                    subject, encoding = decode_header(msg["Subject"])[0]
                    if isinstance(subject, bytes):
                        subject = subject.decode(encoding or "utf-8", errors="ignore")
                    
                    sender = msg.get("From", "")
                    body = ""
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == "text/plain":
                                body = part.get_payload(decode=True).decode(errors="ignore")
                                break
                    else:
                        body = msg.get_payload(decode=True).decode(errors="ignore")
                        
                    subject_lower = subject.lower()
                    body_lower = body.lower()
                    
                    for app in apps:
                        comp = app.get("company", "")
                        role = app.get("role", "")
                        if not comp:
                            continue
                            
                        comp_lower = comp.lower().strip()
                        if comp_lower in subject_lower or comp_lower in sender.lower() or comp_lower in body_lower:
                            current_status = app.get("status", "")
                            new_status = None
                            note_text = f"Email subject: {subject}"
                            
                            if "interview" in subject_lower or "schedule" in subject_lower or "chat with" in subject_lower:
                                new_status = "Interviewing"
                            elif "thank you for applying" in subject_lower or "received your application" in subject_lower or "application confirmation" in subject_lower:
                                if current_status == "Wishlist":
                                    new_status = "Applied"
                            elif "not moving forward" in body_lower or "unfortunately" in body_lower or "other candidates" in body_lower:
                                new_status = "Rejected"
                            elif "offer" in subject_lower or "congratulations" in subject_lower:
                                new_status = "Offer"
                                
                            if new_status and new_status != current_status:
                                app["status"] = new_status
                                curr_notes = app.get("notes", "")
                                app["notes"] = f"{curr_notes} | Auto-updated: {new_status} ({note_text})" if curr_notes else f"Auto-updated: {new_status} ({note_text})"
                                print(f"Auto-updated {role} at {comp} to '{new_status}' based on email: '{subject}'")
                                updates_made = True
                                
        if updates_made:
            save_apps(file_path, apps)
            print("Spreadsheet updated successfully with email scanning matches.")
        else:
            print("No new status updates found from recent emails.")
            
        mail.close()
        mail.logout()
    except Exception as e:
        print(f"IMAP Email connection/sync failed: {str(e)}")

def list_applications_json(file_path):
    apps = load_apps(file_path)
    print(json.dumps(apps))

def main():
    if len(sys.argv) < 3:
        print("Usage:")
        print("  python manage_applications.py <tracker_path> add <company> <role> <location> <link> [resume] [cover_letter] [notes]")
        print("  python manage_applications.py <tracker_path> update <company> <role> <new_status> [notes]")
        print("  python manage_applications.py <tracker_path> delete <company> <role>")
        print("  python manage_applications.py <tracker_path> sync <email> <password> <imap_server>")
        print("  python manage_applications.py <tracker_path> list")
        sys.exit(1)
        
    tracker_path = sys.argv[1]
    
    # If the user passes a .xlsx path, redirect to .json for JSON transition
    if tracker_path.endswith(".xlsx"):
        tracker_path = tracker_path[:-5] + ".json"
        
    action = sys.argv[2].lower()
    
    if action == "list":
        list_applications_json(tracker_path)
    elif action == "add":
        if len(sys.argv) < 7:
            print("Missing arguments for 'add' command.")
            sys.exit(1)
        company = sys.argv[3]
        role = sys.argv[4]
        location = sys.argv[5]
        link = sys.argv[6]
        resume = sys.argv[7] if len(sys.argv) > 7 else ""
        cover_letter = sys.argv[8] if len(sys.argv) > 8 else ""
        notes = sys.argv[9] if len(sys.argv) > 9 else ""
        add_application(tracker_path, company, role, location, link, "Applied", resume, cover_letter, notes)
    elif action == "update":
        if len(sys.argv) < 6:
            print("Missing arguments for 'update' command.")
            sys.exit(1)
        company = sys.argv[3]
        role = sys.argv[4]
        new_status = sys.argv[5]
        notes = sys.argv[6] if len(sys.argv) > 6 else None
        update_application_status(tracker_path, company, role, new_status, notes)
    elif action == "delete":
        if len(sys.argv) < 5:
            print("Missing arguments for 'delete' command.")
            sys.exit(1)
        company = sys.argv[3]
        role = sys.argv[4]
        delete_application(tracker_path, company, role)
    elif action == "sync":
        if len(sys.argv) < 6:
            print("Missing arguments for 'sync' command.")
            sys.exit(1)
        email_addr = sys.argv[3]
        password = sys.argv[4]
        imap_server = sys.argv[5]
        scan_emails_and_sync(tracker_path, email_addr, password, imap_server)
    else:
        print(f"Unknown action '{action}'")
        sys.exit(1)

if __name__ == "__main__":
    main()
